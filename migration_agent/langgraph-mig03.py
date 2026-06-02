"""
LLM-Powered Migration Agent - urllib to requests

This agent uses LangGraph with an LLM (Claude) to intelligently migrate
urllib code to requests. It learns from more real-world examples from the dataset
to understand migration patterns and context.

Architecture:
1. Load a larger set of examples from dataset
2. Create few-shot training prompt
3. Build agent with conditional routing
4. Process user's urllib code through migration pipeline
5. Return optimized requests code
"""

from typing import Annotated, Literal, Optional
from typing_extensions import TypedDict
from langgraph.graph.message import add_messages
from langgraph.graph import StateGraph, START, END
from langchain_core.messages import BaseMessage, HumanMessage, AIMessage, SystemMessage
from langchain_groq import ChatGroq
from langchain_ollama import ChatOllama
import openpyxl
import os
import json
import re
import urllib.request as _urllib_request
from pathlib import Path
from dotenv import load_dotenv


load_dotenv()
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"), override=True)

# ── LLM backend detection (runs once at import time) ─────────────────────────
_OLLAMA_HOST  = os.getenv("OLLAMA_HOST", "http://localhost:11434")
_OLLAMA_MODEL = os.getenv("MIGRATION_OLLAMA_MODEL", os.getenv("REVIEW_OLLAMA_MODEL", "llama3.1:8b"))


def _detectar_ollama() -> tuple[bool, str]:
    """Returns (available, model_name). 2s timeout so missing Ollama never blocks startup."""
    try:
        with _urllib_request.urlopen(f"{_OLLAMA_HOST}/api/tags", timeout=2) as resp:
            data = json.loads(resp.read())
        modelos = [m["name"] for m in data.get("models", [])]
        if not modelos:
            return False, ""
        if _OLLAMA_MODEL in modelos:
            return True, _OLLAMA_MODEL
        match = next((m for m in modelos if m.startswith(_OLLAMA_MODEL.split(":")[0])), None)
        return (True, match) if match else (True, modelos[0])
    except Exception:
        return False, ""


_OLLAMA_DISPONIVEL, _OLLAMA_MODEL_ATIVO = _detectar_ollama()

if _OLLAMA_DISPONIVEL:
    print(f"  [migration_agent] Ollama detected — model: {_OLLAMA_MODEL_ATIVO}")
else:
    print("  [migration_agent] Ollama not detected — using Groq llama-3.3-70b-versatile")
# API_KEY = os.getenv("API_KEY")
PROJECT_ROOT = Path(__file__).resolve().parent.parent
URL_MIGRATE_PATH = PROJECT_ROOT / "url-migrate.py"

# When this file is imported (e.g., by the API gateway), avoid writing files.
WRITE_ARTIFACTS = __name__ == "__main__"


def _criar_modelo_migracao():
    """Create the LLM instance used by migration and refinement nodes."""
    if _OLLAMA_DISPONIVEL:
        return (
            ChatOllama(
                model=_OLLAMA_MODEL_ATIVO,
                base_url=_OLLAMA_HOST,
                temperature=0.0,
            ),
            f"Ollama ({_OLLAMA_MODEL_ATIVO})",
        )

    return (
        ChatGroq(
            model="qwen2.5:14b",
            temperature=0.0,
        ),
        "Groq (qwen2.5:14b)",
    )

# =============================================================================
# DATASET LOADING & TRAINING EXAMPLES
# =============================================================================

def carregar_exemplos_treino(num_exemplos: int = 20) -> list[dict]:
    """
    Load first N examples from dataset for few-shot learning.
    
    Args:
        num_exemplos: Number of examples to load (default 20)
    
    Returns:
        List of example dictionaries with before/after code
    """
    dataset_path = "dataset/Request-Urllib.xlsx"
    
    if not os.path.exists(dataset_path):
        print("⚠️ Dataset not found!")
        return []
    
    try:
        wb = openpyxl.load_workbook(dataset_path)
        ws = wb.active
        
        exemplos = []
        for row_idx in range(2, min(ws.max_row + 1, num_exemplos + 2)):
            repo_name = ws.cell(row_idx, 1).value
            file_name = ws.cell(row_idx, 3).value
            tipo = ws.cell(row_idx, 4).value
            code_before = ws.cell(row_idx, 7).value
            code_after = ws.cell(row_idx, 8).value
            
            if code_before and code_after:
                exemplos.append({
                    "repo": repo_name,
                    "file": file_name,
                    "type": tipo,
                    "before": code_before,
                    "after": code_after
                })
        
        print(f"✅ Loaded {len(exemplos)} training examples from dataset")
        return exemplos
    except Exception as e:
        print(f"❌ Error loading dataset: {e}")
        return []


def criar_prompt_treino(exemplos: list[dict]) -> str:
    """
    Create few-shot prompt with training examples.
    
    Args:
        exemplos: List of training examples
    
    Returns:
        Formatted prompt with examples
    """
    prompt = """You are an expert Python code migration specialist. Your task is to migrate Python code from urllib to requests library.

Key migration rules:
1. IMPORTS: Replace "from urllib.request import ..." with "import requests"
2. METHODS: Replace urlopen() with requests.get() or requests.post()
3. HEADERS: Replace response.getheader('name') with response.headers.get('name')
4. READING: Replace response.read().decode('utf-8') with response.text
5. EXCEPTIONS: Replace urllib.error exceptions with requests.exceptions
6. POST DATA: Convert urllib requests with data parameter to requests.post()
7. TIMEOUT PRESERVATION: Preserve all timeout configurations during migration.
8. REQUEST OBJECT MIGRATION: Convert urllib.request.Request objects into requests API parameters, while preserving the original request behavior.
9. SESSION AND COOKIE HANDLING: When urllib uses cookies, handlers, or openers: migrate to requests.Session(), preserve authentication state and cookies and preserve persistent connections whenever possible.
    Here are the real-world examples from GitHub migrations:

"""
    
    for i, exemplo in enumerate(exemplos, 1):
        prompt += f"\n{'='*80}\nExample {i}: {exemplo['repo']} ({exemplo['type']})\n{'='*80}\n"
        prompt += f"\nBEFORE (urllib):\n```python\n{exemplo['before']}\n```\n"
        prompt += f"\nAFTER (requests):\n```python\n{exemplo['after']}\n```\n"
    
    prompt += f"""

{'='*80}
Now, when you receive urllib code, migrate it following these patterns:

1. Analyze the urllib patterns in the code
2. Map each pattern to the corresponding requests equivalent
3. Preserve the code structure and functionality
4. Handle error cases appropriately
5. Return the migrated Python code with the explanation

Important: Return All valid Python code, but not include news markdown and explanations."""
    
    return prompt


# =============================================================================
# STATE DEFINITION
# =============================================================================

class EstadoAgente(TypedDict):
    """
    State for the migration agent pipeline.
    
    Attributes:
        messages: Conversation history
        codigo_usuario: User's urllib code to migrate
        codigo_migrado: Migrated requests code
        status: Processing status
    """
    messages: Annotated[list[BaseMessage], add_messages]
    codigo_usuario: str
    codigo_migrado: str
    feedback_revisao: str
    status: str


# =============================================================================
# NODES
# =============================================================================

def no_receber_codigo(estado: EstadoAgente) -> dict:
    """
    Receive and validate user's urllib code.
    
    Args:
        estado: Current state
    
    Returns:
        Updated state with code validation
    """
    codigo = estado["codigo_usuario"]
    
    # Check if code contains urllib
    if "urllib" not in codigo:
        return {
            "messages": [AIMessage(content="⚠️ Nenhum código urllib detectado. Por favor, forneça código com urllib.")],
            "status": "no_urllib"
        }
    
    linhas = len(codigo.split('\n'))
    padroes = []
    if "urllib.request" in codigo:
        padroes.append("urllib.request")
    if "urllib2" in codigo:
        padroes.append("urllib2")
    if "urlopen" in codigo:
        padroes.append("urlopen")
    if "urllib.error" in codigo:
        padroes.append("urllib.error")
    
    mensagem = f"📝 Código recebido: {linhas} linhas | Padrões urllib: {', '.join(padroes)}"
    
    return {
        "messages": [AIMessage(content=mensagem)],
        "status": "codigo_recebido"
    }


def no_migrar_com_llm(estado: EstadoAgente, exemplos_treino: list[dict], prompt_sistema: str) -> dict:
    """
    Use LLM (Groq) to intelligently migrate code.
    
    Args:
        estado: Current state
        exemplos_treino: Training examples for few-shot learning
        prompt_sistema: System prompt with training examples
    
    Returns:
        Updated state with migrated code
    """
    codigo_usuario = estado["codigo_usuario"]

    try:
        model, backend_label = _criar_modelo_migracao()
        
        # Create messages
        messages = [
            SystemMessage(content=prompt_sistema),
            HumanMessage(content=f"""Migrate this urllib code to requests:

```python
{codigo_usuario}
```

Return ONLY the migrated Python code without any explanation or markdown.""")
        ]
        
        # Get migration from LLM
        response = model.invoke(messages)
        codigo_migrado = response if isinstance(response, str) else response.content

        # Clean up markdown or prefatory text so only Python source remains
        if "```" in codigo_migrado:
            match = re.search(r"```(?:python|py)?\s*(.*?)\s*```", codigo_migrado, flags=re.DOTALL | re.IGNORECASE)
            if match:
                codigo_migrado = match.group(1).strip()

        if codigo_migrado and not codigo_migrado.lstrip().startswith(("import ", "from ", "def ", "class ", "#", "\"\"\"", "'")):
            linhas = codigo_migrado.splitlines()
            inicio = 0
            for idx, linha in enumerate(linhas):
                texto = linha.lstrip()
                if texto.startswith(("import ", "from ", "def ", "class ", "#", "\"\"\"", "'")):
                    inicio = idx
                    break
            codigo_migrado = "\n".join(linhas[inicio:]).strip()
        
        mensagem = f"🔄 Migração concluída com sucesso usando {backend_label}"
        
        return {
            "messages": [AIMessage(content=mensagem)],
            "codigo_migrado": codigo_migrado,
            "status": "migrado"
        }
    
    except Exception as e:
        erro_msg = f"❌ Erro na migração: {str(e)}"
        return {
            "messages": [AIMessage(content=erro_msg)],
            "status": "erro"
        }


def no_refinar_com_feedback(estado: EstadoAgente, exemplos_treino: list[dict], prompt_sistema: str) -> dict:
    """
    Refine the migrated code using feedback from the review agent.

    Args:
        estado: Current state
        exemplos_treino: Training examples for few-shot learning
        prompt_sistema: System prompt with training examples

    Returns:
        Updated state with refined migrated code
    """
    feedback_revisao = estado.get("feedback_revisao", "").strip()
    codigo_usuario = estado.get("codigo_usuario", "")
    codigo_migrado_atual = estado.get("codigo_migrado", "")

    if not feedback_revisao:
        return {
            "messages": [AIMessage(content="ℹ️ Nenhum feedback do review agent recebido; mantendo a migração atual.")],
            "status": "validado"
        }

    if not codigo_migrado_atual.strip():
        return {
            "messages": [AIMessage(content="⚠️ Feedback recebido, mas não há código migrado para refinar.")],
            "status": "erro"
        }

    try:
        model, backend_label = _criar_modelo_migracao()

        messages = [
            SystemMessage(content=prompt_sistema),
            HumanMessage(content=f"""Refine the existing urllib → requests migration using the review agent feedback.

Original urllib code:
```python
{codigo_usuario}
```

Current migrated code:
```python
{codigo_migrado_atual}
```

Review feedback:
{feedback_revisao}

Return ONLY the improved Python code without any explanation or markdown.""")
        ]

        response = model.invoke(messages)
        codigo_refinado = response if isinstance(response, str) else response.content

        if "```" in codigo_refinado:
            match = re.search(r"```(?:python|py)?\s*(.*?)\s*```", codigo_refinado, flags=re.DOTALL | re.IGNORECASE)
            if match:
                codigo_refinado = match.group(1).strip()

        if codigo_refinado and not codigo_refinado.lstrip().startswith(("import ", "from ", "def ", "class ", "#", "\"\"\"", "'")):
            linhas = codigo_refinado.splitlines()
            inicio = 0
            for idx, linha in enumerate(linhas):
                texto = linha.lstrip()
                if texto.startswith(("import ", "from ", "def ", "class ", "#", "\"\"\"", "'")):
                    inicio = idx
                    break
            codigo_refinado = "\n".join(linhas[inicio:]).strip()

        mensagem = f"🔁 Feedback do review aplicado com sucesso usando {backend_label}"

        return {
            "messages": [AIMessage(content=mensagem)],
            "codigo_migrado": codigo_refinado,
            "status": "feedback_aplicado"
        }

    except Exception as e:
        erro_msg = f"❌ Erro ao aplicar feedback do review: {str(e)}"
        return {
            "messages": [AIMessage(content=erro_msg)],
            "status": "erro"
        }


def no_validar_migracao(estado: EstadoAgente) -> dict:
    """
    Validate the migrated code.
    
    Args:
        estado: Current state
    
    Returns:
        Validation results
    """
    codigo_migrado = estado.get("codigo_migrado", "")

    codigo_lower = codigo_migrado.lower()
    contem_urllib_legado = any(
        token in codigo_lower
        for token in [
            "urllib.request",
            "urllib2",
            "urllib.error",
            "urlopen(",
        ]
    )
    
    validacoes = {
        "Sem urllib legado (request/error/urlopen)": not contem_urllib_legado,
        "Com import requests": "import requests" in codigo_migrado or "requests." in codigo_migrado,
        "Sem urlopen direto": "urlopen(" not in codigo_migrado,
        "Código não vazio": len(codigo_migrado) > 0,
    }
    
    validacoes_ok = sum(1 for v in validacoes.values() if v)
    total = len(validacoes)
    
    resultado = f"🔍 Validação: {validacoes_ok}/{total} critérios atendidos\n"
    for validacao, passou in validacoes.items():
        status = "✓" if passou else "✗"
        resultado += f"  {status} {validacao}\n"
    
    return {
        "messages": [AIMessage(content=resultado)],
        "status": "validado"
    }


# =============================================================================
# CONDITIONAL EDGE
# =============================================================================

def decidir_proxima_etapa(estado: EstadoAgente) -> Literal["inferir", "migrar", "refinar", "validar", "fim"]:
    """
    Decide next step based on current status.
    
    Args:
        estado: Current state
    
    Returns:
        Next node name
    """
    status = estado.get("status", "")
    
    if status == "no_urllib":
        return "fim"
    elif status == "codigo_recebido":
        return "migrar"
    elif status == "migrado":
        if estado.get("feedback_revisao", "").strip():
            return "refinar"
        return "validar"
    elif status == "feedback_aplicado":
        return "validar"
    else:
        return "fim"


# =============================================================================
# GRAPH CONSTRUCTION
# =============================================================================

def criar_agente_migracao(exemplos_treino: list[dict], prompt_sistema: str):
    """
    Build the migration agent graph.
    
    Args:
        exemplos_treino: Training examples
        prompt_sistema: System prompt with examples
    
    Returns:
        Compiled graph
    """
    grafo = StateGraph(EstadoAgente)
    
    # Add nodes
    grafo.add_node("receber", no_receber_codigo)
    grafo.add_node("migrar", lambda estado: no_migrar_com_llm(estado, exemplos_treino, prompt_sistema))
    grafo.add_node("refinar", lambda estado: no_refinar_com_feedback(estado, exemplos_treino, prompt_sistema))
    grafo.add_node("validar", no_validar_migracao)
    
    # Add edges
    grafo.add_edge(START, "receber")
    grafo.add_conditional_edges(
        "receber",
        decidir_proxima_etapa,
        {
            "migrar": "migrar",
            "fim": END
        }
    )
    grafo.add_conditional_edges(
        "migrar",
        decidir_proxima_etapa,
        {
            "refinar": "refinar",
            "validar": "validar",
            "fim": END
        }
    )
    grafo.add_conditional_edges(
        "refinar",
        decidir_proxima_etapa,
        {
            "validar": "validar",
            "fim": END
        }
    )
    grafo.add_edge("validar", END)
    
    return grafo.compile()


# =============================================================================
# MAIN EXECUTION
# =============================================================================

if __name__ == "__main__":
    print("=" * 80)
    print("🤖 AGENTE DE MIGRAÇÃO COM IA - urllib → requests")
    print("=" * 80)
    
    # Load training examples
    print("\n📚 Carregando exemplos de treinamento...")
    exemplos_treino = carregar_exemplos_treino(30)
    
    if not exemplos_treino:
        print("❌ Não foi possível carregar exemplos de treinamento!")
        exit(1)
    
    # Create training prompt
    print("🧠 Criando prompt de treinamento com IA...")
    prompt_sistema = criar_prompt_treino(exemplos_treino)
    
    # Create agent
    print("🔧 Construindo agente de migração...")
    agente = criar_agente_migracao(exemplos_treino, prompt_sistema)

    print("\n" + "=" * 80)
    print("📝 LENDO CÓDIGO DE url.py")
    print("=" * 80)
    
    # Read urllib code from url.py
    try:
        url_file_path = os.path.join(os.path.dirname(__file__), "..", "url.py")
        with open(url_file_path, "r") as f:
            codigo_usuario = f.read().strip()
    except FileNotFoundError:
        print(f"❌ Arquivo url.py não encontrado em {url_file_path}")
        raise SystemExit(1)
    except Exception as e:
        print(f"❌ Erro ao ler url.py: {e}")
        raise SystemExit(1)

    if not codigo_usuario:
        print("❌ Arquivo url.py está vazio. Encerrando.")
        raise SystemExit(1)

    print(f"✅ Código lido com sucesso ({len(codigo_usuario.split(chr(10)))} linhas)")
    print("\n⚙️ PROCESSANDO COM AGENTE IA...\n")
    
    # Run agent
    resultado = agente.invoke({
        "messages": [HumanMessage(content="Migrar código urllib para requests")],
        "codigo_usuario": codigo_usuario,
        "codigo_migrado": "",
        "status": ""
    })
    
    # Display results
    print("\n📊 LOG DE PROCESSAMENTO:")
    print("=" * 80)
    for msg in resultado["messages"]:
        if isinstance(msg, AIMessage):
            print(msg.content)
    
    print("\n" + "=" * 80)
    print("✅ CÓDIGO MIGRADO (REQUESTS):")
    print("=" * 80)
    codigo_final = resultado.get("codigo_migrado", "")
    if not codigo_final.strip():
        print("❌ Migração não gerou código. Arquivo `url-migrate.py` não será sobrescrito.")
        raise SystemExit(1)

    print(codigo_final[:700] + "..." if len(codigo_final) > 700 else codigo_final)

    output_path = URL_MIGRATE_PATH
    try:
        with open(output_path, "w", encoding="utf-8") as arquivo_saida:
            arquivo_saida.write(codigo_final.strip() + "\n")
        print(f"\n💾 Migrated code saved to: {output_path}")
    except Exception as e:
        print(f"❌ Failed to write migrated file: {e}")
    
    print("\n" + "=" * 80)
    print("📚 INFORMAÇÕES DO TREINAMENTO:")
    print("=" * 80)
    print(f"Exemplos carregados: {len(exemplos_treino)}")
    print(f"Modelo: Groq - Llama 3.1 8B Instant")
    print(f"Primeiro exemplo: {exemplos_treino[0]['repo']}")
    print(f"Tipo: {exemplos_treino[0]['type']}")
    print("=" * 80)

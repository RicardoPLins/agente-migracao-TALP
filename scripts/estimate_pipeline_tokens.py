"""Estimativa de tokens por rodada de pipeline (pior caso: max iterações)."""
from __future__ import annotations

import importlib.util
import json
import re
import subprocess
import sys
import tempfile
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent


def tok(text: str | int) -> int:
    return len(str(text)) // 4


def load_review_module():
    spec = importlib.util.spec_from_file_location("ra", REPO / "review_agent" / "review-agent.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def git_diff(original: str, migrated: str) -> str:
    with tempfile.NamedTemporaryFile("w", suffix="_o.py", delete=False, encoding="utf-8") as fo:
        fo.write(original)
        op = fo.name
    with tempfile.NamedTemporaryFile("w", suffix="_m.py", delete=False, encoding="utf-8") as fm:
        fm.write(migrated)
        mp = fm.name
    try:
        r = subprocess.run(["git", "diff", "--no-index", op, mp], capture_output=True, text=True)
        return r.stdout
    finally:
        Path(op).unlink(missing_ok=True)
        Path(mp).unlink(missing_ok=True)


def migration_system_prompt_tokens(num_examples: int = 30) -> int:
    path = REPO / "migration_agent" / "dataset" / "Request-Urllib.xlsx"
    base = 600
    if not path.exists():
        # Dataset ausente: prompt base sem few-shot (~600 tok)
        # Com 30 exemplos típicos do projeto: ~12k–18k tok (use 15k como referência)
        return base if num_examples == 0 else 15_000
    import openpyxl

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ex_chars = 0
    for row_idx in range(2, min(ws.max_row + 1, num_examples + 2)):
        b = ws.cell(row_idx, 7).value or ""
        a = ws.cell(row_idx, 8).value or ""
        if b and a:
            ex_chars += len(str(b)) + len(str(a)) + 120
    return base + tok(ex_chars)


def test_prompt_tokens(name: str, agent_py: str, original: str, migrated: str) -> tuple[int, int]:
    m = re.search(rf'{name} = """(.*?)"""', agent_py, re.S)
    if not m:
        return 0, 0
    base = m.group(1)
    filled = base.replace("{original_code}", original).replace("{migrated_code}", migrated)
    filled = filled.replace("{test_plan}", "{}" * 200).replace("{module_quirks}", "{}" * 200)
    filled = filled.replace("{final_evaluation}", "{}" * 300).replace("{timestamp}", "2026-01-01")
    return tok(base), tok(filled)


def review_prompt_in(key: str, original: str, migrated: str, raw_diff: str, diff_j: dict, trechos: str, achados: str) -> int:
    prompts_dir = REPO / "review_agent" / "prompts"
    regras = (prompts_dir / "regras_migracao.txt").read_text(encoding="utf-8")
    data = json.loads((prompts_dir / f"{key}.json").read_text(encoding="utf-8"))
    tmpl = "\n".join(data["template"]).replace("<<regras_migracao>>", regras)
    t = tok(tmpl)
    diff_str = json.dumps(diff_j, ensure_ascii=False, indent=2)
    extras = {
        "parser": tok(original) + tok(migrated) + tok(raw_diff),
        "classificador": tok(diff_str),
        "agente_semantica": tok(raw_diff) + tok(diff_str) + tok(trechos),
        "agente_seguranca": tok(raw_diff) + tok(diff_str) + tok(trechos),
        "agente_lint_config": tok(original),
        "agente_lint_interpretacao": tok(migrated) + 500,
        "no_critico": tok(achados),
        "relatorio_final": tok(achados) + 20,
    }
    return t + extras.get(key, 0)


def main() -> None:
    original = (REPO / "url.py").read_text(encoding="utf-8")
    migrated_path = REPO / ".pipeline_output" / "migrated_code.py"
    migrated = migrated_path.read_text(encoding="utf-8") if migrated_path.exists() else (
        REPO / "review_agent" / "test1" / "migrado.py"
    ).read_text(encoding="utf-8")

    raw_diff = git_diff(original, migrated)
    ra = load_review_module()
    diff_j = {
        "altered_functions": [
            "build_headers", "encode_query", "build_url", "fetch_users", "fetch_user_by_id",
            "create_user", "update_user", "delete_user", "download_report", "send_form_data",
            "upload_metrics", "fetch_with_retry", "configure_proxy", "ping_service",
            "fetch_binary_asset", "fetch_headers", "submit_feedback", "fetch_secure_data",
            "execute_batch_requests",
        ],
        "added_functions": [],
        "impact_summary": "urllib to requests migration",
    }
    trechos = ra._extrair_trechos_funcoes(migrated, diff_j)
    achados = "\n".join(
        f"- [CONTRACT][P1] `fn{i}` — issue. Trigger: scenario." for i in range(20)
    )

    agent_py = (REPO / "test_agent" / "agent" / "agent.py").read_text(encoding="utf-8")

    # ── Migration (4 runs max) ──
    mig_sys = migration_system_prompt_tokens(30)
    mig_user = tok(original) + 50
    mig_out = tok(migrated)
    mig_per_run = mig_sys + mig_user + mig_out
    mig_runs = 4  # revision_loop 0..3
    mig_total = mig_per_run * mig_runs

    # ── Test agent (4 runs, generator max 3 attempts) ──
    test_calls = {}
    for name in ("PROMPT_ANALYZER", "PROMPT_INSPECTOR", "PROMPT_GENERATOR", "PROMPT_REPORT"):
        _, filled = test_prompt_tokens(name, agent_py, original, migrated)
        test_calls[name] = filled
    gen_out = 2500  # pytest file estimate
    test_per_run = (
        test_calls["PROMPT_ANALYZER"]
        + test_calls["PROMPT_INSPECTOR"]
        + test_calls["PROMPT_GENERATOR"] * 3  # LLM_RETRY_ATTEMPTS
        + gen_out * 3
        + test_calls["PROMPT_REPORT"]
        + 400  # report output
    )
    test_runs = 4
    test_total = test_per_run * test_runs

    # ── Review (3 reflection iterations, all agents) ──
    review_once = {
        "parser": (review_prompt_in("parser", original, migrated, raw_diff, diff_j, trechos, achados), 800),
        "classificador": (review_prompt_in("classificador", original, migrated, raw_diff, diff_j, trechos, achados), 100),
    }
    per_iter = {
        "semantica": (review_prompt_in("agente_semantica", original, migrated, raw_diff, diff_j, trechos, achados), 1500),
        "seguranca": (review_prompt_in("agente_seguranca", original, migrated, raw_diff, diff_j, trechos, achados), 800),
        "lint_config": (review_prompt_in("agente_lint_config", original, migrated, raw_diff, diff_j, trechos, achados), 200),
        "lint_interp": (review_prompt_in("agente_lint_interpretacao", original, migrated, raw_diff, diff_j, trechos, achados), 1200),
        "critico": (review_prompt_in("no_critico", original, migrated, raw_diff, diff_j, trechos, achados), 150),
    }
    relatorio = (
        review_prompt_in("relatorio_final", original, migrated, raw_diff, diff_j, trechos, achados),
        120,
    )

    review_in = sum(v[0] for v in review_once.values()) + 3 * sum(v[0] for v in per_iter.values()) + relatorio[0]
    review_out = sum(v[1] for v in review_once.values()) + 3 * sum(v[1] for v in per_iter.values()) + relatorio[1]
    review_total = review_in + review_out

    grand_in = mig_runs * (mig_sys + mig_user) + test_runs * (
        test_calls["PROMPT_ANALYZER"] + test_calls["PROMPT_INSPECTOR"] + test_calls["PROMPT_GENERATOR"] * 3 + test_calls["PROMPT_REPORT"]
    ) + review_in
    grand_out = mig_runs * mig_out + test_runs * (gen_out * 3 + 400) + review_out
    grand_total = mig_total + test_total + review_total

    print("=== ESTIMATIVA (url.py, pior caso: max iterações) ===")
    print(f"Input original: {len(original)} chars (~{tok(original)} tok)")
    print(f"Input migrated: {len(migrated)} chars (~{tok(migrated)} tok)")
    print(f"raw_diff: ~{tok(raw_diff)} tok | trechos: ~{tok(trechos)} tok")
    print()
    print(f"Migration: {mig_runs} runs × ~{mig_per_run:,} tok = ~{mig_total:,}")
    print(f"  (system+30 few-shot ~{mig_sys:,} + user ~{mig_user:,} + output ~{mig_out:,})")
    print(f"Test: {test_runs} runs × ~{test_per_run:,} tok = ~{test_total:,}")
    print(f"  (generator 3 tentativas/run no pior caso)")
    print(f"Review: 1 run, 3 reflection iters, todos agentes")
    print(f"  input ~{review_in:,} | output ~{review_out:,} | total ~{review_total:,}")
    print()
    print(f"TOTAL RODADA: ~{grand_total:,} tokens (~{grand_in:,} in + ~{grand_out:,} out)")
    print()
    if not (REPO / "migration_agent" / "dataset" / "Request-Urllib.xlsx").exists():
        mig_no_fewshot = mig_runs * (600 + mig_user + mig_out)
        print("NOTA: dataset Request-Urllib.xlsx ausente neste clone.")
        print(f"  Sem few-shot (0 exemplos): migration ~{mig_no_fewshot:,} tok")
        print(f"  Com 30 few-shot (estimativa): migration ~{mig_total:,} tok")
        print(f"  Diferença entre cenários: ~{mig_total - mig_no_fewshot:,} tok")
        print()
    print("Chamadas LLM (pior caso):")
    print(f"  Migration: {mig_runs}")
    print(f"  Test: {test_runs} × (4 + 2 retries generator) = {test_runs * 6} invocações efetivas")
    print(f"  Review: 2 + 3×5 + 1 = 18 invocações")


if __name__ == "__main__":
    main()
